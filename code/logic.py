import os
import sys
from openpyxl import Workbook
import openpyxl
import re

# 定义目录路径
i = 0
if not os.path.exists('code/UUID/uuids'):
    print("这是你第一次使用本程序，请仔细阅读readme文件后使用。")
# 检查目录是否存在
    # 获取用户输入
user_input = input("开始统计？\n请输入 'y' 或 'n': ")

# 使用 if-elif-else 语句检查输入
while i < 1:
    if user_input.lower() == 'y':
        import subprocess
        subprocess.run(["python", "code/UUID/OUTPUT.py"])
        #print("你输入了 '1'。")
        i = 1
    elif user_input.lower() == 'n':
        sys.exit(0)
        i = 1
    else:
         user_input = input("你输入的不是 '1' 也不是 '2'。请重新输入：")

#开始统计
print("开始统计")
# txt文件路径和Excel文件路径
txt_file_path = 'code/UUID/yourfile.txt'
excel_file_path = 'output.xlsx'

# 初始化Excel工作簿和工作表
wb = Workbook()
ws = wb.active
ws.title = 'Contents'

# 读取txt文件内容，从第四行开始
with open(txt_file_path, 'r', encoding='utf-8') as file:
    lines = file.readlines()
    current_line_index = 3  # 从第四行开始
    while current_line_index < len(lines):
        line = lines[current_line_index].strip()

        # 匹配序号及其后的内容
        match = re.match(r'^(\d+)\.\s(.*?)\s(.*)$', line)
        if match:
            # 提取序号和元素
            sequence_number = int(match.group(1))
            element_between_spaces = match.group(2).strip()
            element_after_second_space = match.group(3).strip()

            # 填充到Excel表格中
            # 如果序号对应的行不存在，则创建新行
            if sequence_number > ws.max_row:
                ws.append([])  # 添加新行

            # 将元素填充到对应的列中
            ws.cell(row=sequence_number, column=1, value=element_between_spaces)
            ws.cell(row=sequence_number, column=2, value=element_after_second_space)

            # 序号下一行的所有元素依次填入后续列
            column = 3
            while current_line_index + 1 < len(lines) and not re.match(r'^\d+\.', lines[current_line_index + 1]):
                current_line_index += 1
                line_content = lines[current_line_index].strip()
                ws.cell(row=sequence_number, column=column, value=line_content)
                column += 1
        else:
            # 如果当前行不是序号，则跳过这一行
            current_line_index += 1

        # 保存Excel文件
wb.save(excel_file_path)
print(f"内容已填充到Excel文件: {excel_file_path}")
